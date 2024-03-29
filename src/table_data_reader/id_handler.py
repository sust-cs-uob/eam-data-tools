import openpyxl

from datetime import date
import logging
logger = logging.getLogger(__name__)


def build_id_dict(filename):
    """
    Read through each sheet of the workbook given by the filename
    Build some dictionary mapping variables to their ids
    Find variables with missing ids and flag them and give them a value(?)
    Find variables with duplicate ids and raise an error
    Return the constructed dictionary (and the wb if changed?)

    :param filename: The filename of the excel workbook to be parsed
    :return: A dictionary of Dict: (String -> Dict: (Int -> Int)) to their id values
    """

    logger.info(f'Opening workbook {filename} to build id dict...')
    wb = openpyxl.load_workbook(filename, data_only=True)
    # keys are sheet names, values are dictionaries with row keys and id values
    id_map = {}
    highest_id = -1

    # iterate through the dedicated group-variable sheets and fill out the ids
    for _sheet_name in wb.sheetnames:
        if _sheet_name in ['changes', 'metadata']:
            continue

        sheet = wb[_sheet_name]
        rows = list(sheet.iter_rows())
        id_column = get_id_column_index(filename, sheet)
        id_map[_sheet_name] = {}

        for row in rows[1:]:
            if row[0].value is None:
                continue
            row_index = row[0].row
            id_val = row[id_column].value
            if id_val is not None:
                highest_id = max(id_val, highest_id)
            id_map[_sheet_name][row_index] = id_val

    wb.close()
    logger.info(f'Closed workbook {filename}')
    return id_map, highest_id


def get_id_column_index(filename, sheet):
    """
    :param filename: The workbook name
    :param sheet: The specific sheet to get the id column for (ZERO INDEXED!)
    :return: The id column index
    """
    rows = list(sheet.iter_rows())
    header = [cell.value for cell in rows[0]]
    
    if 'id' not in header:
        raise Exception(f"{filename}, {sheet} has no id column")

    if header.count('id') > 1:
        raise Exception(f"{filename}, {sheet} has multiple id columns")

    id_column = header.index('id')
    return id_column


def check_for_duplicate_ids(id_map) -> bool:
    """
    Checks for duplicate ids in the map
    :param id_map: The id_map generated by build_id_dict()
    :return: True if duplicate non-null id's exist, false otherwise
    """
    used_ids = set()
    for sheet in id_map.keys():
        for row in id_map[sheet].keys():
            current_id = id_map[sheet][row]
            if current_id is not None and current_id in used_ids:
                return True
            used_ids.add(current_id)
    return False


def add_overwrite_msg_to_workbook(workbook, changed_sheet=None, changed_cell=None):
    changes_msg = f'Overwrote empty cell ids in sheet {changed_sheet}, cell {changed_cell}'
    today = date.today()
    date_msg = today.strftime("%d/%m/%Y")

    if 'changes' in workbook.sheetnames:
        sheet = workbook['changes']
        rows = list(sheet.iter_rows())
        new_row = rows[-1][0].row + 1
        sheet.cell(row=new_row, column=1).value = date_msg
        sheet.cell(row=new_row, column=2).value = changes_msg
    else:
        logger.info(f'Creating new sheet for changes in workbook')
        changes = workbook.create_sheet('changes')
        changes.cell(row=1, column=1).value = date_msg
        changes.cell(row=1, column=2).value = changes_msg


def fill_missing_ids(filename, id_map, highest_id):
    """
    Open the workbook, fill in any None ids from the id_map with incremental values, and save it

    :param filename: Name of the worksheet to be edited
    :param id_map: The id_map generated by build_id_dict()
    :param highest_id: The highest id used by any sheet
    :return:
    """

    logger.info(f'Opening workbook {filename} to fill any missing ids...')
    wb = openpyxl.load_workbook(filename, data_only=True)

    for sheet_name in id_map.keys():
        for row in id_map[sheet_name].keys():
            current_id = id_map[sheet_name][row]
            if current_id is None:
                sheet = wb[sheet_name]
                id_column = get_id_column_index(filename, sheet)
                cell = sheet.cell(row=row, column=id_column+1)
                highest_id += 1
                cell.value = highest_id
                logger.info(f'Overwriting cell id value in cell {row}, {id_column+1}')
                add_overwrite_msg_to_workbook(wb, changed_sheet=sheet_name, changed_cell=cell.coordinate)

    wb.save(filename)
    wb.close()
    logger.info(f'Saved and closed workbook {filename}')
